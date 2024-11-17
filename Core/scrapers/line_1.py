import os
import json
from openpyxl import load_workbook
import unicodedata
def normalize_str(s) -> str: 

    return unicodedata.normalize("NFC", s).strip()
def line_1():
    wb2 = load_workbook(os.getcwd() + '/Core/excels/line_1.xlsx')

    ws3 = wb2['تجريش عادي']

    stations = []

    for i in range(3, ws3.max_row, 2):
        if ws3.cell(row=5, column=i).value == None:
            break

        stations.append(normalize_str(ws3.cell(row=5, column=i).value))

    x = {'line_1': {
        'عادی': {'تجريش': {i : [] for i in stations}, 'كهريزك': {i : [] for i in stations}},
        'پنجشنبه': {'تجريش': {i : [] for i in stations}, 'كهريزك': {i : [] for i in stations}},
        'جمعه': {'تجريش': {i : [] for i in stations}, 'كهريزك': {i : [] for i in stations}},
    }}

# Tejrish - Adi
    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=5, column=col).value == None:
            break  # for break if there is no value in a column

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value == None:

                x['line_1']['عادی']["كهريزك"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_1']['عادی']["كهريزك"][normalize_str(ws3.cell(row=5, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

# Tejrish - Panjshanbe
    ws3 = wb2['تجريش پنجشنبه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=5, column=col).value == None:
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value == None:
                x['line_1']['پنجشنبه']["كهريزك"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_1']['پنجشنبه']["كهريزك"][normalize_str(ws3.cell(row=5, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

# Tejrish - Jomeh
    ws3 = wb2['تجريش جمعه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value == None:
                x['line_1']['جمعه']["كهريزك"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_1']['جمعه']["كهريزك"][normalize_str(ws3.cell(row=6, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

# Kahrizak - Adi
    ws3 = wb2['كهريزك عادي']
    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None and ws3.cell(row=6, column=col + 1).value == None:
            break

        for row in range(7, ws3.max_row):
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None and ws3.cell(row=row + 2, column=col).value == None and ws3.cell(row=row + 3, column=col).value == None and ws3.cell(row=row + 4, column=col).value == None and ws3.cell(row=row + 5, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value == None:
                x['line_1']['عادی']["تجريش"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_1']['عادی']["تجريش"][normalize_str(ws3.cell(row=6, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

# Karizak - Panjshanbe
    ws3 = wb2['كهريزك پنجشنبه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(7, ws3.max_row):
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value == None:
                x['line_1']['پنجشنبه']["تجريش"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_1']['پنجشنبه']["تجريش"][normalize_str(ws3.cell(row=6, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

# Karizak - Jomeh
    ws3 = wb2['كهريزك جمعه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(7, ws3.max_row):
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None and ws3.cell(row=row + 2, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value == None:
                x['line_1']['جمعه']["تجريش"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_1']['جمعه']["تجريش"][normalize_str(ws3.cell(row=6, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

    with open(os.getcwd() + '/Core/static/stations_1.json', 'w', encoding='UTF-8') as file:
        file.write(json.dumps(x, ensure_ascii=False))
    return stations
line_1()