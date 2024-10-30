# stable
import os
import json
from openpyxl import load_workbook
import unicodedata

def normalize_str(s):
    return unicodedata.normalize('NFC', s).strip()
def line_2() : 
    wb2 = load_workbook(os.getcwd() + '/Core/static/excels/line_2.xlsx')

    ws3 = wb2['صادقيه عادي']

    stations = []


    for i in range(3, ws3.max_row, 2):
        if ws3.cell(row=5, column=i).value == None:
            break
        stations.append(normalize_str(ws3.cell(row=5, column=i).value))


    x = {'line_2': {

        'عادی': {'فرهنگسرا': {}, 'تهران (صادقيه)': {}},
        'پنجشنبه': {'فرهنگسرا': {}, 'تهران (صادقيه)': {}},
        'جمعه': {'فرهنگسرا': {}, 'تهران (صادقيه)': {}},

    }}

    for j in stations:
        x['line_2']['عادی']["فرهنگسرا"][j] = []
        x['line_2']['پنجشنبه']["فرهنگسرا"][j] = []
        x['line_2']['جمعه']["فرهنگسرا"][j] = []

    for j in stations:
        x['line_2']['عادی']["تهران (صادقيه)"][j] = []
        x['line_2']['پنجشنبه']["تهران (صادقيه)"][j] = []
        x['line_2']['جمعه']["تهران (صادقيه)"][j] = []


    # stable ============
    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=5, column=col).value == None:

            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_2']['عادی']["فرهنگسرا"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_2']['عادی']["فرهنگسرا"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_2']['عادی']["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['صادقيه پنج شنبه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=5, column=col).value == None:
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_2']['پنجشنبه']["فرهنگسرا"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_2']['پنجشنبه']["فرهنگسرا"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_2']['پنجشنبه']["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['صادقيه جمعه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_2']['جمعه']["فرهنگسرا"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_2']['جمعه']["فرهنگسرا"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_2']['جمعه']["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

    # ============
    ws3 = wb2['فرهنگسرا عادي']
    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:

            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_2']['عادی']["تهران (صادقيه)"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_2']['عادی']["تهران (صادقيه)"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_2']['عادی']["تهران (صادقيه)"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['فرهنگسرا پنج شنبه']

    for col in range(3, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:

            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_2']['پنجشنبه']["تهران (صادقيه)"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:

                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_2']['پنجشنبه']["تهران (صادقيه)"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_2']['پنجشنبه']["تهران (صادقيه)"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['فرهنگسرا جمعه']

    for col in range(3, ws3.max_column, 2):
        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_2']['جمعه']["تهران (صادقيه)"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue
            
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None and ws3.cell(row=row + 2, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_2']['جمعه']["تهران (صادقيه)"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_2']['جمعه']["تهران (صادقيه)"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    with open(os.getcwd() + '/Core/static/stations_2.json', 'w', encoding='UTF-8') as file:
        file.write(json.dumps(x, ensure_ascii=False))
    return stations