# stable
import os
import json
from openpyxl import load_workbook
import unicodedata

def normalize_str(s):
    return unicodedata.normalize('NFC', s).strip()
def line_4() : 
    wb2 = load_workbook(os.getcwd() + '/Core/excels/line_4.xlsx')

    ws3 = wb2['علامه - عادي']

    stations = []

    def contains_digit(s):
        return any(char.isdigit() for char in s)

    for i in range(3, ws3.max_column, 1):
    

        if ws3.cell(row=5, column=i).value != None :
            if not contains_digit(str(ws3.cell(row=5, column=i).value)) :
                stations.append(normalize_str(ws3.cell(row=5, column=i).value))
        else :
            if ws3.cell(row=5, column=i).value == None and ws3.cell(row=5, column=i + 1).value == None :
                break




    x = {'line_4': {

        'عادی': {'شهيد كلاهدوز': {}, 'علامه جعفري': {}},
        'پنجشنبه': {'شهيد كلاهدوز': {}, 'علامه جعفري': {}},
        'جمعه': {'شهيد كلاهدوز': {}, 'علامه جعفري': {}},

    }}

    for j in stations:
        x['line_4']['عادی']["شهيد كلاهدوز"][j] = []
        x['line_4']['پنجشنبه']["شهيد كلاهدوز"][j] = []
        x['line_4']['جمعه']["شهيد كلاهدوز"][j] = []

    for j in stations:
        x['line_4']['عادی']["علامه جعفري"][j] = []
        x['line_4']['پنجشنبه']["علامه جعفري"][j] = []
        x['line_4']['جمعه']["علامه جعفري"][j] = []


    # stable ============
    for col in range(3, ws3.max_column, 1):
        
        
        if ws3.cell(row=5, column=col).value == None or contains_digit(str(ws3.cell(row=5, column=col).value)):
            
            continue
        if ws3.cell(row=5, column=col).value == None and ws3.cell(row=5, column=col + 1).value == None :
            
            break
        
        
        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_4']['عادی']["شهيد كلاهدوز"][normalize_str(ws3.cell( row=5, column=col).value)].append("None")
                continue

            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == None:
                x['line_4']['عادی']["شهيد كلاهدوز"][normalize_str(ws3.cell( row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            
            x['line_4']['عادی']["شهيد كلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append( f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")
                



    ws3 = wb2['علامه - پنجشنبه']

    for col in range(3, ws3.max_column, 1):
        

        if ws3.cell(row=5, column=col).value == None or contains_digit(str(ws3.cell(row=5, column=col).value)):
            continue
        if ws3.cell(row=5, column=col).value == None and ws3.cell(row=5, column=col + 1).value == None :
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_4']['پنجشنبه']["شهيد كلاهدوز"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_4']['پنجشنبه']["شهيد كلاهدوز"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_4']['پنجشنبه']["شهيد كلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['علامه - جمعه']

    for col in range(3, ws3.max_column, 1):
        

        if ws3.cell(row=5, column=col).value == None or contains_digit(str(ws3.cell(row=5, column=col).value)):
            continue
        if ws3.cell(row=5, column=col).value == None and ws3.cell(row=5, column=col + 1).value == None :
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_4']['جمعه']["شهيد كلاهدوز"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_4']['جمعه']["شهيد كلاهدوز"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_4']['جمعه']["شهيد كلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

    # ============
    ws3 = wb2['شهيد كلاهدوز - عادي']
    for col in range(3, ws3.max_column, 1):
        

        if ws3.cell(row=5, column=col).value == None or contains_digit(str(ws3.cell(row=5, column=col).value)):
            continue
        if ws3.cell(row=5, column=col).value == None and ws3.cell(row=5, column=col + 1).value == None :
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_4']['عادی']["علامه جعفري"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_4']['عادی']["علامه جعفري"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_4']['عادی']["علامه جعفري"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['شهيد كلاهدوز - پنجشنبه']

    for col in range(3, ws3.max_column, 1):
        

        if ws3.cell(row=5, column=col).value == None or contains_digit(str(ws3.cell(row=5, column=col).value)):
            continue
        if ws3.cell(row=5, column=col).value == None and ws3.cell(row=5, column=col + 1).value == None :
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_4']['پنجشنبه']["علامه جعفري"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:

                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_4']['پنجشنبه']["علامه جعفري"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_4']['پنجشنبه']["علامه جعفري"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['شهيد كلاهدوز - جمعه']

    for col in range(3, ws3.max_column, 1):
        

        if ws3.cell(row=5, column=col).value == None or contains_digit(str(ws3.cell(row=5, column=col).value)):
            continue
        if ws3.cell(row=5, column=col).value == None and ws3.cell(row=5, column=col + 1).value == None :
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_4']['جمعه']["علامه جعفري"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue
            
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None and ws3.cell(row=row + 2, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_4']['جمعه']["علامه جعفري"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_4']['جمعه']["علامه جعفري"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    with open(os.getcwd() + '/Core/static/stations_4.json', 'w', encoding='UTF-8') as file:
        file.write(json.dumps(x, ensure_ascii=False))
    return stations