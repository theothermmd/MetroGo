# stable
import os
import json
from openpyxl import load_workbook
import unicodedata

def normalize_str(s):
    return unicodedata.normalize('NFC', s).strip()
def line_5() : 
    wb2 = load_workbook(os.getcwd() + '/Core/excels/line_5.xlsx')
    ws3 = wb2['گلشهر- عادي']

    stations = []


    for i in range(4, ws3.max_column, 1):
        if ws3.cell(row=5, column=i).value == None:
            break
        stations.append(normalize_str(ws3.cell(row=5, column=i).value))


    x = {'line_5': {

        'عادی': {'صادقيه': {}, 'گلشهر': {}},
        'پنجشنبه': {'صادقيه': {}, 'گلشهر': {}},
        'جمعه': {'صادقيه': {}, 'گلشهر': {}},

    }}

    for j in stations:
        x['line_5']['عادی']["صادقيه"][j] = []
        x['line_5']['پنجشنبه']["صادقيه"][j] = []
        x['line_5']['جمعه']["صادقيه"][j] = []

    for j in stations:
        x['line_5']['عادی']["گلشهر"][j] = []
        x['line_5']['پنجشنبه']["گلشهر"][j] = []
        x['line_5']['جمعه']["گلشهر"][j] = []


    # stable ============
    for col in range(4, ws3.max_column, 1):

        if ws3.cell(row=5, column=col).value == None:

            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_5']['عادی']["صادقيه"][normalize_str(ws3.cell( row=5, column=col).value)].append("None")
                continue

            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == None:
                x['line_5']['عادی']["صادقيه"][normalize_str(ws3.cell( row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            
            x['line_5']['عادی']["صادقيه"][normalize_str(ws3.cell(row=5, column=col).value)].append( f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")
                




    ws3 = wb2['گلشهر- پنجشنبه']

    for col in range(4, ws3.max_column, 1):

        if ws3.cell(row=5, column=col).value == None:
            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_5']['پنجشنبه']["صادقيه"][normalize_str(ws3.cell( row=5, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_5']['پنجشنبه']["صادقيه"][normalize_str(ws3.cell( row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_5']['پنجشنبه']["صادقيه"][normalize_str(ws3.cell(row=5, column=col).value)].append( f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['گلشهر - جمعه و تعطيلات']

    for col in range(4, ws3.max_column, 1):

        if ws3.cell(row=5, column=col).value == None:
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_5']['جمعه']["صادقيه"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_5']['جمعه']["صادقيه"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_5']['جمعه']["صادقيه"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

    # ============
    ws3 = wb2['صادقيه - عادي']
    for col in range(4, ws3.max_column, 1):

        if ws3.cell(row=5, column=col).value == None:

            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_5']['عادی']["گلشهر"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_5']['عادی']["گلشهر"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_5']['عادی']["گلشهر"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['صادقيه - پنجشنبه']

    for col in range(4, ws3.max_column, 1):

        if ws3.cell(row=5, column=col).value == None:

            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_5']['پنجشنبه']["گلشهر"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:

                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_5']['پنجشنبه']["گلشهر"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_5']['پنجشنبه']["گلشهر"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['صادقيه - تعطيل']

    for col in range(4, ws3.max_column, 1):
        if ws3.cell(row=5, column=col).value == None:
            break

        for row in range(6, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_5']['جمعه']["گلشهر"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue
            
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None and ws3.cell(row=row + 2, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_5']['جمعه']["گلشهر"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_5']['جمعه']["گلشهر"][normalize_str(ws3.cell(row=5, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")



    with open(os.getcwd() + '/Core/static/stations_5.json', 'w', encoding='UTF-8') as file:
        file.write(json.dumps(x, ensure_ascii=False))
    return stations