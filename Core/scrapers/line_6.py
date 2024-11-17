# stable
import os
import json
from openpyxl import load_workbook
import unicodedata

def normalize_str(s):
    return unicodedata.normalize('NFC', s).strip()
def line_6() : 
    wb2 = load_workbook(os.getcwd() + '/Core/excels/line_6.xlsx')

    ws3 = wb2['شهيدآرمان - عادي']

    stations = []


    for i in range(4, ws3.max_column, 2):
        if ws3.cell(row=6, column=i).value == None:
            break
        stations.append(normalize_str(ws3.cell(row=6, column=i).value))


    x = {'line_6': {

        'عادی': {'دولت آباد': {}, 'شهيد آرمان': {}},
        'تعطيل': {'دولت آباد': {}, 'شهيد آرمان': {}},

    }}

    for j in stations:
        x['line_6']['عادی']["دولت آباد"][j] = []
        x['line_6']['تعطيل']["دولت آباد"][j] = []


    for j in stations:
        x['line_6']['عادی']["شهيد آرمان"][j] = []
        x['line_6']['تعطيل']["شهيد آرمان"][j] = []



    # stable ============
    for col in range(4, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:

            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_6']['عادی']["دولت آباد"][normalize_str(ws3.cell( row=6, column=col).value)].append("None")
                continue

            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break

            if ws3.cell(row=row, column=col).value == None:
                x['line_6']['عادی']["دولت آباد"][normalize_str(ws3.cell( row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            
            x['line_6']['عادی']["دولت آباد"][normalize_str(ws3.cell(row=6, column=col).value)].append( f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")
                




    ws3 = wb2['شهيدآرمان - تعطيل']

    for col in range(4, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_6']['تعطيل']["دولت آباد"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_6']['تعطيل']["دولت آباد"][normalize_str(ws3.cell(
                    row=5, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_6']['تعطيل']["دولت آباد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")


    ws3 = wb2['دولت آباد - عادي']

    for col in range(4, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:
            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_6']['عادی']["شهيد آرمان"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_6']['عادی']["شهيد آرمان"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value
            x['line_6']['عادی']["شهيد آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

    # ============
    ws3 = wb2['دولت آباد - تعطيل']
    for col in range(4, ws3.max_column, 2):

        if ws3.cell(row=6, column=col).value == None:

            break

        for row in range(7, ws3.max_row):

            if ws3.cell(row=row, column=col).value == "":
                x['line_6']['تعطيل']["شهيد آرمان"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")

                continue
            if ws3.cell(row=row, column=col).value == None and ws3.cell(row=row + 1, column=col).value == None:
                break
            if ws3.cell(row=row, column=col).value == None:
                x['line_6']['تعطيل']["شهيد آرمان"][normalize_str(ws3.cell(
                    row=6, column=col).value)].append("None")
                continue

            k = ws3.cell(row=row, column=col).value

            x['line_6']['تعطيل']["شهيد آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
                f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")



    with open(os.getcwd() + '/Core/static/stations_6.json', 'w', encoding='UTF-8') as file:
        file.write(json.dumps(x, ensure_ascii=False))
    return stations