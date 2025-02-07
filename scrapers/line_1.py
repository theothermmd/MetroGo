import os
import json
from openpyxl import load_workbook


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_1():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_1.xlsx"))

	ws3 = wb2["تجريش عادي"]

	stations = []

	for i in range(3, ws3.max_row + 1, 2):
		if ws3.cell(row=5, column=i).value is None:
			break

		stations.append(normalize_str(ws3.cell(row=5, column=i).value))

	x = {
		"line_1": {
			"عادی": {
				"تجریش": {i: [] for i in stations},
				"کهریزک": {i: [] for i in stations},
			},
			"پنجشنبه": {
				"تجریش": {i: [] for i in stations},
				"کهریزک": {i: [] for i in stations},
			},
			"جمعه": {
				"تجریش": {i: [] for i in stations},
				"کهریزک": {i: [] for i in stations},
			},
		}
	}

	# Tejrish - Adi
	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=5, column=col).value is None:
			break  # for break if there is no value in a column

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_1"]["عادی"]["کهریزک"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_1"]["عادی"]["کهریزک"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# Tejrish - Panjshanbe
	ws3 = wb2["تجريش پنجشنبه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_1"]["پنجشنبه"]["کهریزک"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_1"]["پنجشنبه"]["کهریزک"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# Tejrish - Jomeh
	ws3 = wb2["تجريش جمعه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_1"]["جمعه"]["کهریزک"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_1"]["جمعه"]["کهریزک"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# Kahrizak - Adi
	ws3 = wb2["كهريزك عادي"]
	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None and ws3.cell(row=6, column=col + 1).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if (
				ws3.cell(row=row, column=col).value is None
				and ws3.cell(row=row + 1, column=col).value is None
				and ws3.cell(row=row + 2, column=col).value is None
				and ws3.cell(row=row + 3, column=col).value is None
				and ws3.cell(row=row + 4, column=col).value is None
				and ws3.cell(row=row + 5, column=col).value is None
			):
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_1"]["عادی"]["تجریش"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_1"]["عادی"]["تجریش"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# Karizak - Panjshanbe
	ws3 = wb2["كهريزك پنجشنبه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_1"]["پنجشنبه"]["تجریش"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_1"]["پنجشنبه"]["تجریش"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# Karizak - Jomeh
	ws3 = wb2["كهريزك جمعه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if (
				ws3.cell(row=row, column=col).value is None
				and ws3.cell(row=row + 1, column=col).value is None
				and ws3.cell(row=row + 2, column=col).value is None
			):
				break
			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_1"]["جمعه"]["تجریش"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_1"]["جمعه"]["تجریش"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")
	os.makedirs(base_path, exist_ok=True)
	with open(os.path.join(base_path, "line_1.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))

	return stations
