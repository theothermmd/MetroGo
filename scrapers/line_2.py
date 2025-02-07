# stable
import os
import json
from openpyxl import load_workbook


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_2():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_2.xlsx"))

	ws3 = wb2["صادقيه عادي"]

	stations = []

	for i in range(3, ws3.max_row + 1, 2):
		if ws3.cell(row=5, column=i).value is None:
			break
		stations.append(normalize_str(ws3.cell(row=5, column=i).value))

	x = {
		"line_2": {
			"عادی": {
				"فرهنگسرا": {i: [] for i in stations},
				"تهران (صادقیه)": {i: [] for i in stations},
			},
			"پنجشنبه": {
				"فرهنگسرا": {i: [] for i in stations},
				"تهران (صادقیه)": {i: [] for i in stations},
			},
			"جمعه": {
				"فرهنگسرا": {i: [] for i in stations},
				"تهران (صادقیه)": {i: [] for i in stations},
			},
		}
	}

	# stable ============
	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_2"]["عادی"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_2"]["عادی"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_2"]["عادی"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["صادقيه پنج شنبه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_2"]["پنجشنبه"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_2"]["پنجشنبه"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_2"]["پنجشنبه"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["صادقيه جمعه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_2"]["جمعه"]["فرهنگسرا"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_2"]["جمعه"]["فرهنگسرا"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_2"]["جمعه"]["فرهنگسرا"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# ============
	ws3 = wb2["فرهنگسرا عادي"]
	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_2"]["عادی"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_2"]["عادی"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_2"]["عادی"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["فرهنگسرا پنج شنبه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_2"]["پنجشنبه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_2"]["پنجشنبه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_2"]["پنجشنبه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["فرهنگسرا جمعه"]

	for col in range(3, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_2"]["جمعه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			if (
				ws3.cell(row=row, column=col).value is None
				and ws3.cell(row=row + 1, column=col).value is None
				and ws3.cell(row=row + 2, column=col).value is None
			):
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_2"]["جمعه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_2"]["جمعه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")

	with open(os.path.join(base_path, "line_2.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))
	return stations
