# stable
import os
import json
from openpyxl import load_workbook


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_3():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_3.xlsx"))

	ws3 = wb2["آزادگان - عادی"]

	stations = []
	for i in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=i).value is None:
			break
		stations.append(normalize_str(ws3.cell(row=6, column=i).value))
	# input(stations)

	x = {
		"line_3": {
			"عادی": {"قائم": {}, "آزادگان": {}},
			"پنجشنبه": {"قائم": {}, "آزادگان": {}},
			"جمعه": {"قائم": {}, "آزادگان": {}},
		}
	}

	for j in stations:
		x["line_3"]["عادی"]["قائم"][j] = []
		x["line_3"]["پنجشنبه"]["قائم"][j] = []
		x["line_3"]["جمعه"]["قائم"][j] = []

	for j in stations:
		x["line_3"]["عادی"]["آزادگان"][j] = []
		x["line_3"]["پنجشنبه"]["آزادگان"][j] = []
		x["line_3"]["جمعه"]["آزادگان"][j] = []

	# stable ============
	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_3"]["عادی"]["قائم"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value is None:
				x["line_3"]["عادی"]["قائم"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_3"]["عادی"]["قائم"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["آزادگان - پنجشنبه"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_3"]["پنجشنبه"]["قائم"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_3"]["پنجشنبه"]["قائم"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_3"]["پنجشنبه"]["قائم"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["آزادگان - جمعه"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_3"]["جمعه"]["قائم"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_3"]["جمعه"]["قائم"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_3"]["جمعه"]["قائم"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# ============
	ws3 = wb2["قائم - عادی"]
	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_3"]["عادی"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_3"]["عادی"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_3"]["عادی"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["قائم - پنجشنبه"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_3"]["پنجشنبه"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_3"]["پنجشنبه"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_3"]["پنجشنبه"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["قائم - جمعه"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_3"]["جمعه"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			if (
				ws3.cell(row=row, column=col).value is None
				and ws3.cell(row=row + 1, column=col).value is None
				and ws3.cell(row=row + 2, column=col).value is None
			):
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_3"]["جمعه"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_3"]["جمعه"]["آزادگان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")

	with open(os.path.join(base_path, "line_3.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))
	return stations


line_3()
