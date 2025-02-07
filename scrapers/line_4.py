# stable
import os
import json
from openpyxl import load_workbook
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_4():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_4.xlsx"))

	ws3 = wb2["علامه - عادي"]

	stations = []

	def contains_digit(s):
		return any(char.isdigit() for char in s)

	for i in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=i).value is not None:
			if not contains_digit(str(ws3.cell(row=5, column=i).value)):
				stations.append(normalize_str(ws3.cell(row=5, column=i).value))
		else:
			if ws3.cell(row=5, column=i).value is None and ws3.cell(row=5, column=i + 1).value is None:
				break

	x = {
		"line_4": {
			"عادی": {"شهید کلاهدوز": {}, "علامه جعفری": {}},
			"پنجشنبه": {"شهید کلاهدوز": {}, "علامه جعفری": {}},
			"جمعه": {"شهید کلاهدوز": {}, "علامه جعفری": {}},
		}
	}

	for j in stations:
		x["line_4"]["عادی"]["شهید کلاهدوز"][j] = []
		x["line_4"]["پنجشنبه"]["شهید کلاهدوز"][j] = []
		x["line_4"]["جمعه"]["شهید کلاهدوز"][j] = []

	for j in stations:
		x["line_4"]["عادی"]["علامه جعفری"][j] = []
		x["line_4"]["پنجشنبه"]["علامه جعفری"][j] = []
		x["line_4"]["جمعه"]["علامه جعفری"][j] = []

	# stable ============
	for col in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None or contains_digit(str(ws3.cell(row=5, column=col).value)):
			continue
		if ws3.cell(row=5, column=col).value is None and ws3.cell(row=5, column=col + 1).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_4"]["عادی"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value is None:
				x["line_4"]["عادی"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_4"]["عادی"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["علامه - پنجشنبه"]

	for col in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None or contains_digit(str(ws3.cell(row=5, column=col).value)):
			continue
		if ws3.cell(row=5, column=col).value is None and ws3.cell(row=5, column=col + 1).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_4"]["پنجشنبه"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_4"]["پنجشنبه"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_4"]["پنجشنبه"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["علامه - جمعه"]

	for col in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None or contains_digit(str(ws3.cell(row=5, column=col).value)):
			continue
		if ws3.cell(row=5, column=col).value is None and ws3.cell(row=5, column=col + 1).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_4"]["جمعه"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_4"]["جمعه"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_4"]["جمعه"]["شهید کلاهدوز"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# ============
	ws3 = wb2["شهيد كلاهدوز - عادي"]
	for col in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None or contains_digit(str(ws3.cell(row=5, column=col).value)):
			continue
		if ws3.cell(row=5, column=col).value is None and ws3.cell(row=5, column=col + 1).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_4"]["عادی"]["علامه جعفری"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_4"]["عادی"]["علامه جعفری"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_4"]["عادی"]["علامه جعفری"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["شهيد كلاهدوز - پنجشنبه"]

	for col in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None or contains_digit(str(ws3.cell(row=5, column=col).value)):
			continue
		if ws3.cell(row=5, column=col).value is None and ws3.cell(row=5, column=col + 1).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_4"]["پنجشنبه"]["علامه جعفری"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_4"]["پنجشنبه"]["علامه جعفری"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_4"]["پنجشنبه"]["علامه جعفری"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["شهيد كلاهدوز - جمعه"]

	for col in range(3, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None or contains_digit(str(ws3.cell(row=5, column=col).value)):
			continue
		if ws3.cell(row=5, column=col).value is None and ws3.cell(row=5, column=col + 1).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_4"]["جمعه"]["علامه جعفری"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			if (
				ws3.cell(row=row, column=col).value is None
				and ws3.cell(row=row + 1, column=col).value is None
				and ws3.cell(row=row + 2, column=col).value is None
			):
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_4"]["جمعه"]["علامه جعفری"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_4"]["جمعه"]["علامه جعفری"][normalize_str(ws3.cell(row=5, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")

	with open(os.path.join(base_path, "line_4.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))
	return stations
