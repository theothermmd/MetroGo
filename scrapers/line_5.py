# stable
import os
import json
from openpyxl import load_workbook


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_5():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_5.xlsx"))
	ws3 = wb2["گلشهر- عادي"]

	stations = []

	for i in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=i).value is None:
			break
		if normalize_str(ws3.cell(row=5, column=i).value) == "صادقیه":
			stations.append("تهران (صادقیه)")
		else:
			stations.append(normalize_str(ws3.cell(row=5, column=i).value))

	x = {
		"line_5": {
			"عادی": {"تهران (صادقیه)": {}, "گلشهر": {}},
			"پنجشنبه": {"تهران (صادقیه)": {}, "گلشهر": {}},
			"جمعه": {"تهران (صادقیه)": {}, "گلشهر": {}},
		}
	}

	for j in stations:
		x["line_5"]["عادی"]["تهران (صادقیه)"][j] = []
		x["line_5"]["پنجشنبه"]["تهران (صادقیه)"][j] = []
		x["line_5"]["جمعه"]["تهران (صادقیه)"][j] = []

	for j in stations:
		x["line_5"]["عادی"]["گلشهر"][j] = []
		x["line_5"]["پنجشنبه"]["گلشهر"][j] = []
		x["line_5"]["جمعه"]["گلشهر"][j] = []

	# stable ============
	for col in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_5"]["عادی"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value is None:
				x["line_5"]["عادی"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			station = normalize_str(ws3.cell(row=5, column=col).value)
			if station == "صادقیه":
				station = "تهران (صادقیه)"
			x["line_5"]["عادی"]["تهران (صادقیه)"][station].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["گلشهر- پنجشنبه"]

	for col in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_5"]["پنجشنبه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_5"]["پنجشنبه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			station = normalize_str(ws3.cell(row=5, column=col).value)
			if station == "صادقیه":
				station = "تهران (صادقیه)"
			x["line_5"]["پنجشنبه"]["تهران (صادقیه)"][station].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["گلشهر - جمعه و تعطيلات"]

	for col in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_5"]["جمعه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_5"]["جمعه"]["تهران (صادقیه)"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			station = normalize_str(ws3.cell(row=5, column=col).value)
			if station == "صادقیه":
				station = "تهران (صادقیه)"
			x["line_5"]["جمعه"]["تهران (صادقیه)"][station].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# ============
	ws3 = wb2["صادقيه - عادي"]
	for col in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_5"]["عادی"]["گلشهر"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_5"]["عادی"]["گلشهر"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			station = normalize_str(ws3.cell(row=5, column=col).value)
			if station == "صادقیه":
				station = "تهران (صادقیه)"
			x["line_5"]["عادی"]["گلشهر"][station].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

	ws3 = wb2["صادقيه - پنجشنبه"]

	for col in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_5"]["پنجشنبه"]["گلشهر"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_5"]["پنجشنبه"]["گلشهر"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			station = normalize_str(ws3.cell(row=5, column=col).value)
			if station == "صادقیه":
				station = "تهران (صادقیه)"
			x["line_5"]["پنجشنبه"]["گلشهر"][station].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

	ws3 = wb2["صادقيه - تعطيل"]

	for col in range(4, ws3.max_column, 1):
		if ws3.cell(row=5, column=col).value is None:
			break

		for row in range(6, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_5"]["جمعه"]["گلشهر"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			if (
				ws3.cell(row=row, column=col).value is None
				and ws3.cell(row=row + 1, column=col).value is None
				and ws3.cell(row=row + 2, column=col).value is None
			):
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_5"]["جمعه"]["گلشهر"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			station = normalize_str(ws3.cell(row=5, column=col).value)
			if station == "صادقیه":
				station = "تهران (صادقیه)"
			x["line_5"]["جمعه"]["گلشهر"][station].append(f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}")

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")

	with open(os.path.join(base_path, "line_5.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))
	return stations
