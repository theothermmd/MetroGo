# stable
import os
import json
from openpyxl import load_workbook


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_6():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_6.xlsx"))

	ws3 = wb2["شهيدآرمان - عادي"]

	stations = []

	for i in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=i).value is None:
			break
		stations.append(normalize_str(ws3.cell(row=6, column=i).value))

	x = {
		"line_6": {
			"عادی": {"دولت آباد": {}, "شهید آرمان": {}},
			"تعطيل": {"دولت آباد": {}, "شهید آرمان": {}},
		}
	}

	for j in stations:
		x["line_6"]["عادی"]["دولت آباد"][j] = []
		x["line_6"]["تعطيل"]["دولت آباد"][j] = []

	for j in stations:
		x["line_6"]["عادی"]["شهید آرمان"][j] = []
		x["line_6"]["تعطيل"]["شهید آرمان"][j] = []

	# stable ============
	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_6"]["عادی"]["دولت آباد"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value is None:
				x["line_6"]["عادی"]["دولت آباد"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_6"]["عادی"]["دولت آباد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["شهيدآرمان - تعطيل"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_6"]["تعطيل"]["دولت آباد"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_6"]["تعطيل"]["دولت آباد"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_6"]["تعطيل"]["دولت آباد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["دولت آباد - عادي"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_6"]["عادی"]["شهید آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_6"]["عادی"]["شهید آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_6"]["عادی"]["شهید آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# ============
	ws3 = wb2["دولت آباد - تعطيل"]
	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_6"]["تعطيل"]["شهید آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_6"]["تعطيل"]["شهید آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_6"]["تعطيل"]["شهید آرمان"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")
	x["line_6"]["پنجشنبه"] = x["line_6"]["عادی"]
	x["line_6"]["جمعه"] = x["line_6"]["تعطيل"]
	x["line_6"].pop("تعطيل")
	with open(os.path.join(base_path, "line_6.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))

	return stations
