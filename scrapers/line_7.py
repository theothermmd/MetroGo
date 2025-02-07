# stable
import os
import json
from openpyxl import load_workbook


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(s)


def line_7():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "line_7.xlsx"))

	ws3 = wb2["كتاب - عادي"]

	stations = []

	for i in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=i).value is None:
			break
		stations.append(normalize_str(ws3.cell(row=6, column=i).value))

	x = {
		"line_7": {
			"عادی": {"بسیج": {}, "میدان کتاب": {}},
			"تعطيل": {"بسیج": {}, "میدان کتاب": {}},
		}
	}

	for j in stations:
		x["line_7"]["عادی"]["بسیج"][normalize_str(j)] = []
		x["line_7"]["تعطيل"]["بسیج"][normalize_str(j)] = []

	for j in stations:
		x["line_7"]["عادی"]["میدان کتاب"][normalize_str(j)] = []
		x["line_7"]["تعطيل"]["میدان کتاب"][normalize_str(j)] = []

	# stable ============
	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_7"]["عادی"]["بسیج"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value is None:
				x["line_7"]["عادی"]["بسیج"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_7"]["عادی"]["بسیج"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["كتاب - تعطیل"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_7"]["تعطيل"]["بسیج"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_7"]["تعطيل"]["بسیج"][normalize_str(ws3.cell(row=5, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_7"]["تعطيل"]["بسیج"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["بسيج - عادي"]

	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_7"]["عادی"]["میدان کتاب"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_7"]["عادی"]["میدان کتاب"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value
			x["line_7"]["عادی"]["میدان کتاب"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	# ============
	ws3 = wb2["بسيج - تعطيل"]
	for col in range(4, ws3.max_column, 2):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value == "":
				x["line_7"]["تعطيل"]["میدان کتاب"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")

				continue
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break
			if ws3.cell(row=row, column=col).value is None:
				x["line_7"]["تعطيل"]["میدان کتاب"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_7"]["تعطيل"]["میدان کتاب"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")
	x["line_7"]["پنجشنبه"] = x["line_7"]["عادی"]
	x["line_7"]["جمعه"] = x["line_7"]["تعطيل"]
	x["line_7"].pop("تعطيل")
	with open(os.path.join(base_path, "line_7.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))
	return stations
