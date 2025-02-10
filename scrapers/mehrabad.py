import os
import json
from openpyxl import load_workbook
import unicodedata


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str_s(s) -> str:
	return unicodedata.normalize("NFC", s).strip()


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(normalize_str_s(s))


def mehrabad():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "mehrabad.xlsx"))

	ws3 = wb2["بيمه عادي"]

	stations = []

	for i in range(4, ws3.max_row + 1, 2):
		if ws3.cell(row=6, column=i).value is None:
			break

		stations.append(normalize_str(ws3.cell(row=6, column=i).value))


	x = {"line_mehrabad":
	  {
		"عادی" : {
			"بیمه": {i: [] for i in stations},
		    "مهرآباد": {i: [] for i in stations}
			},
		"پنجشنبه" : {
			"بیمه": {i: [] for i in stations},
		    "مهرآباد": {i: [] for i in stations}
			},
		"جمعه" : {
			"بیمه": {i: [] for i in stations},
		    "مهرآباد": {i: [] for i in stations}
			}

		  }

		}

	for col in range(4, 10, 2):

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_mehrabad"]['عادی']["مهرآباد"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_mehrabad"]['عادی']["مهرآباد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	for col in range(4, 10, 2):

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_mehrabad"]['پنجشنبه']["مهرآباد"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_mehrabad"]['پنجشنبه']["مهرآباد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)


	ws3 = wb2["بيمه تعطيل"]

	for col in range(4, 10, 2):
		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_mehrabad"]['جمعه']["مهرآباد"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_mehrabad"]['جمعه']["مهرآباد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)



	ws3 = wb2["مهرآباد- عادي"]

	for col in range(5, 10, 2):

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_mehrabad"]['عادی']["بیمه"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_mehrabad"]['عادی']["بیمه"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	for col in range(5, 10, 2):

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_mehrabad"]['پنجشنبه']["بیمه"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_mehrabad"]['پنجشنبه']["بیمه"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)


	ws3 = wb2["مهرآباد تعطيل"]

	for col in range(5, 10, 2):

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["line_mehrabad"]['جمعه']["بیمه"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["line_mehrabad"]['جمعه']["بیمه"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)




	base_path = os.path.join(os.getcwd(), "assets", "time_lines")

	with open(os.path.join(base_path, "line_mehrabad.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(x, ensure_ascii=False))

	return stations

mehrabad()
