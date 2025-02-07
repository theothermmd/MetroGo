import os
import json
from openpyxl import load_workbook
import unicodedata


import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import libs.WordUtils as WordUtils


def normalize_str_s(s) -> str:
	return unicodedata.normalize("NFC", s).strip()


def normalize_str(s):
	return WordUtils.WordUtils.correct_persian_text(normalize_str_s(s))


def parand():
	base_path = os.path.join(os.getcwd(), "assets", "excels")
	wb2 = load_workbook(os.path.join(base_path, "parand.xlsx"))

	ws3 = wb2["پرند"]

	stations = []

	for i in range(2, ws3.max_row + 1, 1):
		if ws3.cell(row=6, column=i).value is None:
			break

		stations.append(normalize_str(ws3.cell(row=6, column=i).value))

	x = {"parand": {"شاهد": {i: [] for i in stations}, "پرند": {i: [] for i in stations}}}

	for col in range(2, ws3.max_column, 1):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["parand"]["پرند"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["parand"]["پرند"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	ws3 = wb2["شاهد"]

	for col in range(2, ws3.max_column, 1):
		if ws3.cell(row=6, column=col).value is None:
			break

		for row in range(7, ws3.max_row + 1):
			if ws3.cell(row=row, column=col).value is None and ws3.cell(row=row + 1, column=col).value is None:
				break

			if ws3.cell(row=row, column=col).value == "" or ws3.cell(row=row, column=col).value is None:
				x["parand"]["شاهد"][normalize_str(ws3.cell(row=6, column=col).value)].append("None")
				continue

			k = ws3.cell(row=row, column=col).value

			x["parand"]["شاهد"][normalize_str(ws3.cell(row=6, column=col).value)].append(
				f"{k.hour}:{'0' + str(k.minute) if k.minute < 10 else k.minute}"
			)

	base_path = os.path.join(os.getcwd(), "assets", "time_lines")
	s = {"line_parand": {"عادی": x["parand"], "پنجشنبه": x["parand"], "جمعه": x["parand"]}}
	with open(os.path.join(base_path, "line_parand.json"), "w", encoding="UTF-8") as file:
		file.write(json.dumps(s, ensure_ascii=False))

	return stations
